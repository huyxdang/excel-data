"""
BRD Sheet Converter - Excel Sheet to Markdown-ready CSV

Converts Excel worksheets to CSV format with cell coordinates and markdown image
references. Designed for LLM processing pipelines that need structured data with
embedded image links.

Why this exists:
    Excel stores images separately from cell data in its internal XML structure.
    The openpyxl library doesn't reliably load images, so this module directly
    parses the underlying XML within the .xlsx ZIP archive to extract:
    - Image positions (which cell they're anchored to)
    - Image files (extracted to a local folder)
    - Markdown references (for rendering in documentation)

Excel .xlsx structure:
    .xlsx files are ZIP archives containing:
    ├── xl/
    │   ├── worksheets/sheet1.xml    # Cell data (text, numbers, formulas)
    │   ├── drawings/drawing1.xml    # Image anchor positions
    │   ├── drawings/_rels/*.rels    # Maps rId -> image filename
    │   └── media/image1.png         # Actual image files

Usage:
    python sheet_converter.py <excel_file> [output_csv] [sheet_name]

Example:
    python sheet_converter.py input.xlsx output.csv "5.1.2a"

Output:
    output.csv              # CSV with markdown image references
    images/                 # Extracted image files
        B6_Picture_2.png
"""

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from zipfile import ZipFile
import xml.etree.ElementTree as ET
import sys
import os
import re
from typing import Optional, Dict, List


# XML namespaces used in Office Open XML (OOXML) drawing files
NAMESPACES = {
    'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
    'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
    'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
    'rel': 'http://schemas.openxmlformats.org/package/2006/relationships',
}


def get_sheet_drawing_path(excel_path: str, sheet_name: str) -> Optional[str]:
    """
    Find the drawing XML path associated with a specific worksheet.
    
    Each Excel sheet can have an associated drawing file that contains
    image position information. This function looks up the relationship
    between the sheet and its drawing file.
    
    Args:
        excel_path: Path to the .xlsx file
        sheet_name: Name of the worksheet to look up
        
    Returns:
        Path to the drawing XML file within the ZIP archive
        (e.g., 'xl/drawings/drawing2.xml'), or None if no drawing exists
        
    Example:
        >>> get_sheet_drawing_path('report.xlsx', 'Sheet1')
        'xl/drawings/drawing1.xml'
    """
    wb = load_workbook(excel_path, data_only=False)
    sheet = wb[sheet_name]
    drawing_path = None
    
    if hasattr(sheet, '_rels') and sheet._rels:
        for rel in sheet._rels:
            if 'drawing' in rel.Target.lower():
                drawing_path = rel.Target
                # Convert relative path to absolute path within ZIP
                if drawing_path.startswith('..'):
                    drawing_path = 'xl' + drawing_path[2:]
                elif not drawing_path.startswith('xl/'):
                    drawing_path = 'xl/drawings/' + drawing_path.split('/')[-1]
                break
    
    wb.close()
    return drawing_path


def get_drawing_rels(excel_path: str, drawing_path: str) -> Dict[str, str]:
    """
    Parse the drawing relationships file to map rId references to image paths.
    
    Drawing XML files reference images by rId (e.g., 'rId1'). The .rels file
    maps these IDs to actual file paths within the ZIP archive.
    
    Args:
        excel_path: Path to the .xlsx file
        drawing_path: Path to the drawing XML within the ZIP
                      (e.g., 'xl/drawings/drawing2.xml')
        
    Returns:
        Dictionary mapping rId to image path
        Example: {'rId1': 'xl/media/image2.png', 'rId2': 'xl/media/image3.png'}
    """
    rels = {}
    
    # Construct rels path: xl/drawings/drawing2.xml -> xl/drawings/_rels/drawing2.xml.rels
    drawing_filename = drawing_path.split('/')[-1]
    rels_path = f"xl/drawings/_rels/{drawing_filename}.rels"
    
    with ZipFile(excel_path, 'r') as zf:
        if rels_path not in zf.namelist():
            return rels
        
        rels_xml = zf.read(rels_path).decode('utf-8')
        root = ET.fromstring(rels_xml)
        
        for rel in root.findall('.//{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
            rid = rel.get('Id')
            target = rel.get('Target')
            # Convert relative path: ../media/image1.png -> xl/media/image1.png
            if target.startswith('..'):
                target = 'xl' + target[2:]
            rels[rid] = target
    
    return rels


def process_anchor(
    anchor: ET.Element,
    rels: Dict[str, str],
    zf: ZipFile,
    images_dir: str,
    sheet_name: str
) -> Optional[Dict]:
    """
    Process a single image anchor element and extract the image file.
    
    Excel anchors images to cells using either:
    - twoCellAnchor: Image spans from one cell to another
    - oneCellAnchor: Image is anchored to a single cell
    
    Both types have a <from> element specifying the top-left cell position.
    
    Args:
        anchor: XML Element representing the anchor (twoCellAnchor or oneCellAnchor)
        rels: Dictionary mapping rId to image paths (from get_drawing_rels)
        zf: Open ZipFile object for the Excel file
        images_dir: Directory path where extracted images should be saved
        sheet_name: Name of the worksheet (used for unique filenames)
        
    Returns:
        Dictionary with image info if successful:
        {
            'cell': 'B6',
            'filename': '5.1.2a_B6_img001.png',
            'markdown': '![5.1.2a_B6](images/5.1.2a_B6_img001.png)',
            'description': '5.1.2a_B6'
        }
        Returns None if image extraction fails.
    """
    # Get anchor position (top-left cell)
    from_elem = anchor.find('xdr:from', NAMESPACES)
    if from_elem is None:
        return None
    
    # Excel uses 0-indexed col/row in XML, convert to 1-indexed cell reference
    col = int(from_elem.find('xdr:col', NAMESPACES).text)
    row = int(from_elem.find('xdr:row', NAMESPACES).text)
    cell_ref = f"{get_column_letter(col + 1)}{row + 1}"
    
    # Find the picture element within the anchor
    pic = anchor.find('.//xdr:pic', NAMESPACES)
    if pic is None:
        return None
    
    # Get rId reference to find actual image file
    # The blip element contains the embedded image reference
    blip = pic.find('.//a:blip', NAMESPACES)
    if blip is None:
        return None
    
    rid = blip.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed')
    if rid not in rels:
        return None
    
    # Get source image path in ZIP and extract it
    source_path = rels[rid]
    
    # Create unique filename: {sheet}_{cell}_{original_image_name}
    # e.g., "5.1.2a_B6_image2.png"
    safe_sheet = re.sub(r'[^\w\-]', '_', sheet_name)
    original_img_name = os.path.basename(source_path)  # e.g., "image2.png"
    output_filename = f"{safe_sheet}_{cell_ref}_{original_img_name}"
    output_path = os.path.join(images_dir, output_filename)
    
    # Extract image file from ZIP to output directory
    if source_path in zf.namelist():
        with open(output_path, 'wb') as f:
            f.write(zf.read(source_path))
    else:
        return None
    
    # Use sheet_cell as description for clarity in markdown
    description = f"{sheet_name}_{cell_ref}"
    
    return {
        'cell': cell_ref,
        'filename': output_filename,
        'markdown': f"![{description}](images/{output_filename})",
        'description': description
    }


def extract_images_from_drawing(
    excel_path: str,
    drawing_path: str,
    output_dir: str,
    sheet_name: str
) -> List[Dict]:
    """
    Extract all images from a drawing XML file.
    
    Parses the drawing XML to find all image anchors (twoCellAnchor and
    oneCellAnchor elements), extracts the referenced images, and saves
    them to the output directory.
    
    Args:
        excel_path: Path to the .xlsx file
        drawing_path: Path to the drawing XML within the ZIP
        output_dir: Base directory for output (images saved to output_dir/images/)
        sheet_name: Name of the worksheet (used for unique filenames)
        
    Returns:
        List of image info dictionaries (see process_anchor for structure)
    """
    images = []
    
    if not drawing_path:
        return images
    
    # Get rId -> image path mapping
    rels = get_drawing_rels(excel_path, drawing_path)
    if not rels:
        return images
    
    with ZipFile(excel_path, 'r') as zf:
        if drawing_path not in zf.namelist():
            return images
        
        # Parse drawing XML
        drawing_xml = zf.read(drawing_path).decode('utf-8')
        root = ET.fromstring(drawing_xml)
        
        # Create images output directory
        images_dir = os.path.join(output_dir, 'images')
        os.makedirs(images_dir, exist_ok=True)
        
        # Process both anchor types
        for anchor in root.findall('.//xdr:twoCellAnchor', NAMESPACES):
            img_info = process_anchor(anchor, rels, zf, images_dir, sheet_name)
            if img_info:
                images.append(img_info)
        
        for anchor in root.findall('.//xdr:oneCellAnchor', NAMESPACES):
            img_info = process_anchor(anchor, rels, zf, images_dir, sheet_name)
            if img_info:
                images.append(img_info)
    
    return images


def extract_images(excel_path: str, sheet_name: str, output_dir: str) -> List[Dict]:
    """
    Extract all images from a specific worksheet.
    
    High-level function that combines sheet lookup and image extraction.
    
    Args:
        excel_path: Path to the .xlsx file
        sheet_name: Name of the worksheet
        output_dir: Base directory for output
        
    Returns:
        List of image info dictionaries
    """
    drawing_path = get_sheet_drawing_path(excel_path, sheet_name)
    return extract_images_from_drawing(excel_path, drawing_path, output_dir, sheet_name)


def excel_to_csv(excel_path: str, sheet_name: str = None, output_dir: str = '.') -> str:
    """
    Convert an Excel worksheet to CSV with cell coordinates and markdown image refs.
    
    Each cell is output as "COORD: VALUE" format. Images are referenced using
    markdown syntax that will render when the CSV content is processed.
    
    Args:
        excel_path: Path to the .xlsx file
        sheet_name: Name of sheet to convert (None = first/active sheet)
        output_dir: Directory to save extracted images
        
    Returns:
        CSV string with format:
            Sheet: SheetName
            A1: value1,B1: value2,C1: ![Image](images/C1_Picture.png)
            A2: value3,B2:,C2: value4
            
    Note:
        Images don't count toward Excel's max_row/max_col, so this function
        extends the iteration range to include cells containing images.
    """
    # Load workbook to get sheet info
    wb = load_workbook(excel_path, data_only=True)
    
    if sheet_name:
        sheet = wb[sheet_name]
    else:
        sheet = wb.active
        sheet_name = sheet.title
    
    wb.close()
    
    # Extract images using drawing XML (bypasses openpyxl's broken image loading)
    images = extract_images(excel_path, sheet_name, output_dir)
    image_dict = {img['cell']: img['markdown'] for img in images}
    
    # Reload for data extraction
    wb = load_workbook(excel_path, data_only=True)
    sheet = wb[sheet_name]
    
    # Calculate max row/col including image positions
    # (images don't contribute to sheet.max_row/max_col)
    max_row = sheet.max_row
    max_col = sheet.max_column
    
    for img in images:
        cell_ref = img['cell']
        match = re.match(r'([A-Z]+)(\d+)', cell_ref)
        if match:
            img_col = column_index_from_string(match.group(1))
            img_row = int(match.group(2))
            max_row = max(max_row, img_row)
            max_col = max(max_col, img_col)
    
    # Build CSV lines
    csv_lines = [f"Sheet: {sheet.title}"]
    
    for row in sheet.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        row_values = []
        for cell in row:
            col_letter = get_column_letter(cell.column)
            row_num = cell.row
            cell_coord = f"{col_letter}{row_num}"
            
            if cell_coord in image_dict:
                row_values.append(f"{cell_coord}: {image_dict[cell_coord]}")
            elif cell.value is not None:
                value = str(cell.value)
                # Cell contains: "Line 1
                #                 Line 2"
                # value is now: "Line 1\nLine 2" (with actual newline character)
                
                value = value.replace('\n', '\\n')
                # value is now: "Line 1\\nLine 2" 
                # The \n (actual newline) becomes \\n (backslash + letter n)
                # In the file it will display as: Line 1\nLine 2
                
                row_values.append(f"{cell_coord}: {value}")
            else:
                row_values.append(f"{cell_coord}:")
        
        csv_lines.append(','.join(row_values))
    
    wb.close()
    return '\n'.join(csv_lines)


def convert_file(
    excel_path: str,
    output_path: str = None,
    sheet_name: str = None
) -> str:
    """
    Convert an Excel file to CSV with extracted images.
    
    Main entry point for the converter. Handles file I/O and provides
    user feedback about output locations.
    
    Args:
        excel_path: Path to the .xlsx file
        output_path: Path to save CSV (None = print to stdout)
        sheet_name: Sheet to convert (None = first sheet)
        
    Returns:
        The generated CSV string
        
    Side effects:
        - Creates CSV file at output_path (if specified)
        - Creates images/ directory with extracted images
    """
    output_dir = os.path.dirname(output_path) or '.' if output_path else '.'
    
    csv_string = excel_to_csv(excel_path, sheet_name, output_dir)
    
    if output_path:
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(csv_string)
        print(f"Saved CSV to: {output_path}")
        print(f"Images saved to: {os.path.join(output_dir, 'images')}/")
    else:
        print(csv_string)
    
    return csv_string


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    
    excel_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None
    sheet = sys.argv[3] if len(sys.argv) > 3 else None
    
    convert_file(excel_file, output_file, sheet)