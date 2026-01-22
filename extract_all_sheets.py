"""
Extract All Sheets from Excel

Detects all sheets in an Excel file and calls sheet_converter.py 
for each sheet individually.

Usage:
    python extract_all_sheets.py <excel_file> <output_dir>

Example:
    python extract_all_sheets.py input.xlsx output/sheets/
    
Output:
    output/sheets/
    ├── 0.csv
    ├── 1.csv
    ├── 5.csv
    ├── 5.1.1a.csv
    ├── 5.1.1b.csv
    └── ...
    output/images/
    ├── 5.1.1a_B5_image1.png
    └── ...
"""

import sys
import os
import re
import subprocess
from openpyxl import load_workbook


def sanitize_filename(sheet_name: str) -> str:
    """
    Convert sheet name to safe filename.
    
    Examples:
        '5.1.1a' -> '5.1.1a'
        '5.1.1b (2)' -> '5.1.1b_2'
        'Status' -> 'Status'
    """
    safe_name = sheet_name.replace(' ', '_')
    safe_name = safe_name.replace('(', '').replace(')', '')
    safe_name = re.sub(r'[^\w\.\-]', '_', safe_name)
    safe_name = re.sub(r'_+', '_', safe_name)
    safe_name = safe_name.strip('_')
    return safe_name


def get_sheet_names(excel_path: str) -> list:
    """Get all sheet names from Excel file."""
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    wb.close()
    return sheet_names


def extract_all_sheets(excel_path: str, output_dir: str) -> dict:
    """
    Extract all sheets from Excel file.
    
    Args:
        excel_path: Path to Excel file
        output_dir: Directory to save CSV files
        
    Returns:
        Dictionary with extraction results
    """
    # Create output directories
    sheets_dir = os.path.join(output_dir, 'sheets')
    os.makedirs(sheets_dir, exist_ok=True)
    
    # Get script directory (where sheet_converter.py is)
    script_dir = os.path.dirname(os.path.abspath(__file__))
    converter_path = os.path.join(script_dir, 'sheet_converter.py')
    
    if not os.path.exists(converter_path):
        print(f"Error: sheet_converter.py not found at {converter_path}")
        sys.exit(1)
    
    # Get all sheet names
    sheet_names = get_sheet_names(excel_path)
    print(f"Found {len(sheet_names)} sheets: {sheet_names}")
    print("-" * 60)
    
    results = {
        'success': [],
        'failed': [],
        'sheets': {}
    }
    
    for i, sheet_name in enumerate(sheet_names):
        safe_name = sanitize_filename(sheet_name)
        output_csv = os.path.join(sheets_dir, f"{safe_name}.csv")
        
        print(f"[{i+1}/{len(sheet_names)}] Extracting '{sheet_name}'...", end=" ")
        
        try:
            # Call sheet_converter.py with sheet name
            # python sheet_converter.py <excel_file> <output_csv> <sheet_name>
            result = subprocess.run(
                [sys.executable, converter_path, excel_path, output_csv, sheet_name],
                capture_output=True,
                text=True,
                timeout=60
            )
            
            if result.returncode == 0:
                # Get file size
                file_size = os.path.getsize(output_csv) if os.path.exists(output_csv) else 0
                print(f"✓ ({file_size:,} bytes)")
                
                results['success'].append(sheet_name)
                results['sheets'][sheet_name] = {
                    'filename': f"{safe_name}.csv",
                    'path': output_csv,
                    'size': file_size
                }
            else:
                print(f"✗ Error")
                if result.stderr:
                    print(f"    {result.stderr.strip()}")
                results['failed'].append(sheet_name)
                
        except subprocess.TimeoutExpired:
            print(f"✗ Timeout")
            results['failed'].append(sheet_name)
        except Exception as e:
            print(f"✗ {e}")
            results['failed'].append(sheet_name)
    
    # Write manifest
    manifest_path = os.path.join(output_dir, 'manifest.txt')
    with open(manifest_path, 'w', encoding='utf-8') as f:
        f.write(f"# Excel Sheet Extraction Manifest\n")
        f.write(f"# Source: {excel_path}\n")
        f.write(f"# Total sheets: {len(sheet_names)}\n")
        f.write(f"# Successful: {len(results['success'])}\n")
        f.write(f"# Failed: {len(results['failed'])}\n\n")
        
        f.write("## Sheets\n")
        for sheet_name, info in results['sheets'].items():
            f.write(f"{sheet_name}\t{info['filename']}\t{info['size']} bytes\n")
        
        if results['failed']:
            f.write("\n## Failed\n")
            for sheet_name in results['failed']:
                f.write(f"{sheet_name}\n")
    
    return results


def main():
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)
    
    excel_file = sys.argv[1]
    output_dir = sys.argv[2]
    
    if not os.path.exists(excel_file):
        print(f"Error: Excel file not found: {excel_file}")
        sys.exit(1)
    
    print(f"Excel file: {excel_file}")
    print(f"Output dir: {output_dir}")
    print("=" * 60)
    
    results = extract_all_sheets(excel_file, output_dir)
    
    print("=" * 60)
    print(f"✅ Extraction complete")
    print(f"   Success: {len(results['success'])} sheets")
    print(f"   Failed:  {len(results['failed'])} sheets")
    print(f"   Output:  {output_dir}/sheets/")
    
    if results['failed']:
        print(f"\n⚠️  Failed sheets: {results['failed']}")


if __name__ == "__main__":
    main()