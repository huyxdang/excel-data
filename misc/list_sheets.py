"""
List all sheet names in an Excel file
"""

from openpyxl import load_workbook
import sys


def list_sheets(excel_path):
    """
    Display all sheet names in an Excel file
    
    Args:
        excel_path: Path to Excel file
    
    Returns:
        List of sheet names
    """
    wb = load_workbook(excel_path)
    sheets = wb.sheetnames
    
    print(f"Sheets in '{excel_path}':\n")
    for i, sheet_name in enumerate(sheets, 1):
        print(f"  {i}. {sheet_name}")
    
    wb.close()
    return sheets


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python list_sheets.py <excel_file>")
        print("\nExample:")
        print("  python list_sheets.py input.xlsx")
        sys.exit(1)
    
    excel_file = sys.argv[1]
    list_sheets(excel_file)