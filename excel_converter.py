"""
BRD Converter - Converts an entire Excel workbook to a single CSV file

With all sheets processed sequentially

Each sheet's images are extracted
and all content is aggregated into one final output.

Usage:
    python excel_converter.py <excel_file> [output_csv]

Example:
    python excel_converter.py BRD_input.xlsx final_output.csv

Output:
    final_output.csv        # Combined CSV with all sheets
    images/                 # All extracted images from all sheets
        5_1_2a_B6_image2.png
        5_2_1a_C10_image3.png
        ...
"""

from openpyxl import load_workbook
import sys
import os
from typing import List

from utils.sheet_converter import excel_to_csv

def get_all_sheet_names(excel_path: str) -> List[str]:
    """
    Get all sheet names from an Excel workbook in order.
    
    Args:
        excel_path: Path to the .xlsx file
        
    Returns:
        List of sheet names in their workbook order
    """
    wb = load_workbook(excel_path, read_only=True)
    sheet_names = wb.sheetnames
    wb.close()
    return sheet_names


def process_all_sheets(excel_path: str, output_path: str = None) -> str:
    """
    Process all sheets in an Excel workbook and combine into one CSV.
    
    Each sheet is converted to CSV format with cell coordinates and
    markdown image references. All sheets are combined sequentially
    with clear separators between them.
    
    Args:
        excel_path: Path to the .xlsx file
        output_path: Path to save combined CSV (None = print to stdout)
        
    Returns:
        Combined CSV string with all sheets
        
    Output format:
        ================================================================================
        Sheet: 5.1.1a
        ================================================================================
        A1: value1,B1: value2
        A2: value3,B2: ![img](images/5_1_1a_B2_image1.png)
        
        ================================================================================
        Sheet: 5.1.2a
        ================================================================================
        A1: value4,B1: value5
        ...
    """
    output_dir = os.path.dirname(output_path) or '.' if output_path else '.'
    
    # Get all sheet names
    sheet_names = get_all_sheet_names(excel_path)
    print(f"Found {len(sheet_names)} sheets: {sheet_names}")
    
    # Process each sheet
    all_csv_parts = []
    
    for i, sheet_name in enumerate(sheet_names):
        print(f"\nProcessing sheet {i+1}/{len(sheet_names)}: '{sheet_name}'")
        
        try:
            # Convert single sheet to CSV
            csv_content = excel_to_csv(excel_path, sheet_name, output_dir)
            
            # Add separator header for clarity
            separator = "=" * 80
            header = f"\n{separator}\nSheet: {sheet_name}\n{separator}\n"
            
            # Remove the "Sheet: X" line from csv_content since we have a header
            # The first line is always "Sheet: X"
            csv_lines = csv_content.split('\n')
            if csv_lines[0].startswith('Sheet:'):
                csv_lines = csv_lines[1:]
            csv_content = '\n'.join(csv_lines)
            
            all_csv_parts.append(header + csv_content)
            print(f"  ✓ Processed '{sheet_name}'")
            
        except Exception as e:
            print(f"  ✗ Error processing '{sheet_name}': {e}")
            # Add error note to output
            separator = "=" * 80
            error_section = f"\n{separator}\nSheet: {sheet_name} (ERROR: {e})\n{separator}\n"
            all_csv_parts.append(error_section)
    
    # Combine all parts
    combined_csv = '\n'.join(all_csv_parts)
    
    # Save or print
    if output_path:
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(combined_csv)
        print(f"\n{'='*80}")
        print(f"Saved combined CSV to: {output_path}")
        print(f"Images saved to: {os.path.join(output_dir, 'images')}/")
        print(f"Total sheets processed: {len(sheet_names)}")
    else:
        print(combined_csv)
    
    return combined_csv


def process_selected_sheets(
    excel_path: str,
    sheet_names: List[str],
    output_path: str = None
) -> str:
    """
    Process only selected sheets from an Excel workbook.
    
    Useful when you only need specific sheets, not the entire workbook.
    
    Args:
        excel_path: Path to the .xlsx file
        sheet_names: List of sheet names to process (in desired order)
        output_path: Path to save combined CSV (None = print to stdout)
        
    Returns:
        Combined CSV string with selected sheets
    """
    output_dir = os.path.dirname(output_path) or '.' if output_path else '.'
    
    # Validate sheet names
    available_sheets = get_all_sheet_names(excel_path)
    invalid_sheets = [s for s in sheet_names if s not in available_sheets]
    if invalid_sheets:
        print(f"Warning: These sheets don't exist: {invalid_sheets}")
        sheet_names = [s for s in sheet_names if s in available_sheets]
    
    print(f"Processing {len(sheet_names)} selected sheets")
    
    all_csv_parts = []
    
    for i, sheet_name in enumerate(sheet_names):
        print(f"\nProcessing sheet {i+1}/{len(sheet_names)}: '{sheet_name}'")
        
        try:
            csv_content = excel_to_csv(excel_path, sheet_name, output_dir)
            
            separator = "=" * 80
            header = f"\n{separator}\nSheet: {sheet_name}\n{separator}\n"
            
            csv_lines = csv_content.split('\n')
            if csv_lines[0].startswith('Sheet:'):
                csv_lines = csv_lines[1:]
            csv_content = '\n'.join(csv_lines)
            
            all_csv_parts.append(header + csv_content)
            print(f"  ✓ Processed '{sheet_name}'")
            
        except Exception as e:
            print(f"  ✗ Error processing '{sheet_name}': {e}")
            separator = "=" * 80
            error_section = f"\n{separator}\nSheet: {sheet_name} (ERROR: {e})\n{separator}\n"
            all_csv_parts.append(error_section)
    
    combined_csv = '\n'.join(all_csv_parts)
    
    if output_path:
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(combined_csv)
        print(f"\n{'='*80}")
        print(f"Saved combined CSV to: {output_path}")
        print(f"Images saved to: {os.path.join(output_dir, 'images')}/")
    else:
        print(combined_csv)
    
    return combined_csv


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        print("\nAvailable commands:")
        print("  python excel_converter.py <excel_file> [output_csv]")
        print("      Process all sheets")
        print("\n  python excel_converter.py <excel_file> [output_csv] --sheets 'Sheet1,Sheet2,Sheet3'")
        print("      Process only specified sheets (comma-separated)")
        print("\n  python excel_converter.py <excel_file> --list")
        print("      List all sheet names without processing")
        sys.exit(1)
    
    excel_file = sys.argv[1]
    
    # Check for --list flag
    if '--list' in sys.argv:
        sheets = get_all_sheet_names(excel_file)
        print(f"Sheets in '{excel_file}':")
        for i, name in enumerate(sheets):
            print(f"  {i+1}. {name}")
        sys.exit(0)
    
    # Check for --sheets flag
    selected_sheets = None
    if '--sheets' in sys.argv:
        idx = sys.argv.index('--sheets')
        if idx + 1 < len(sys.argv):
            selected_sheets = [s.strip() for s in sys.argv[idx + 1].split(',')]
            # Remove --sheets and its argument from argv for output_file detection
            sys.argv = sys.argv[:idx] + sys.argv[idx+2:]
    
    output_file = sys.argv[2] if len(sys.argv) > 2 else None
    
    if selected_sheets:
        process_selected_sheets(excel_file, selected_sheets, output_file)
    else:
        process_all_sheets(excel_file, output_file)