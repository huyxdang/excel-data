"""
Simple BRD Converter
Converts Excel to CSV format for LLM processing
"""

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
import sys


def extract_images(sheet):
    """Extract all images from the sheet with their cell locations"""
    images = []
    
    if hasattr(sheet, '_images'):
        for img in sheet._images:
            # Get anchor position
            if hasattr(img, 'anchor') and hasattr(img.anchor, '_from'):
                cell_ref = f"{img.anchor._from.col}{img.anchor._from.row}"
                images.append({
                    'cell': cell_ref,
                    'description': f"[IMAGE at {cell_ref}]"
                })
    
    return images


def excel_to_csv(excel_path, sheet_name=None):
    """
    Convert Excel sheet to CSV string
    
    Args:
        excel_path: Path to Excel file
        sheet_name: Name of sheet to convert (None = first sheet)
    
    Returns:
        CSV string with newline separators
    """
    # Load workbook
    wb = load_workbook(excel_path, data_only=True)
    
    # Select sheet
    if sheet_name:
        sheet = wb[sheet_name]
    else:
        sheet = wb.active
    
    # Extract images
    images = extract_images(sheet)
    image_dict = {img['cell']: img['description'] for img in images}
    
    # Convert to CSV
    csv_lines = []
    
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
        row_values = []
        for cell in row:
            # Check if cell has an image
            cell_ref = cell.coordinate
            if cell_ref in image_dict:
                row_values.append(image_dict[cell_ref])
            # Get cell value
            elif cell.value is not None:
                # Convert to string and escape commas
                value = str(cell.value).replace(',', '\\,')
                row_values.append(value)
            else:
                # Empty cell
                row_values.append("")
        
        # Join row with commas
        csv_lines.append(','.join(row_values))
    
    wb.close()
    
    # Join all lines with newlines
    return '\n'.join(csv_lines)


def convert_file(excel_path, output_path=None, sheet_name=None):
    """
    Convert Excel file to CSV
    
    Args:
        excel_path: Path to Excel file
        output_path: Path to save CSV (None = print to stdout)
        sheet_name: Sheet to convert (None = first sheet)
    """
    csv_string = excel_to_csv(excel_path, sheet_name)
    
    if output_path:
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(csv_string)
        print(f"Saved CSV to: {output_path}")
    else:
        print(csv_string)
    
    return csv_string


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python brd_converter_simple.py <excel_file> [output_csv] [sheet_name]")
        print("\nExample:")
        print("  python brd_converter_simple.py input.xlsx output.csv")
        print("  python brd_converter_simple.py input.xlsx output.csv 'Sheet1'")
        sys.exit(1)
    
    excel_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None
    sheet = sys.argv[3] if len(sys.argv) > 3 else None
    
    convert_file(excel_file, output_file, sheet)
