"""
Accuracy & Latency Measurement for BRD Converter

Runs the full excel_converter pipeline from scratch, measuring:
- Processing latency (total and per-sheet)
- Cell content accuracy (comparing output CSV against original Excel)
- Image detection accuracy

Usage:
    python measure.py <excel_file> [output_csv]

Example:
    python measure.py BRD_input.xlsx final_output.csv
"""

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import re
import sys
import os
from typing import Dict, List
from dataclasses import dataclass, field
import time

# Import the converter we're testing
from excel_converter import process_all_sheets, get_all_sheet_names


@dataclass
class SheetMetrics:
    """Metrics for a single sheet."""
    sheet_name: str
    total_cells: int = 0
    matched_cells: int = 0
    mismatched_cells: int = 0
    images_found: int = 0
    mismatches: List[Dict] = field(default_factory=list)
    
    @property
    def accuracy(self) -> float:
        if self.total_cells == 0:
            return 100.0
        return (self.matched_cells / self.total_cells) * 100


@dataclass 
class OverallMetrics:
    """Aggregate metrics across all sheets."""
    sheets: List[SheetMetrics] = field(default_factory=list)
    conversion_time: float = 0.0
    accuracy_check_time: float = 0.0
    
    @property
    def total_cells(self) -> int:
        return sum(s.total_cells for s in self.sheets)
    
    @property
    def total_matched(self) -> int:
        return sum(s.matched_cells for s in self.sheets)
    
    @property
    def total_images(self) -> int:
        return sum(s.images_found for s in self.sheets)
    
    @property
    def overall_accuracy(self) -> float:
        if self.total_cells == 0:
            return 100.0
        return (self.total_matched / self.total_cells) * 100


def parse_csv_output(csv_content: str) -> Dict[str, Dict[str, str]]:
    """
    Parse the combined CSV output into a structured format.
    
    Args:
        csv_content: Raw CSV string with all sheets
        
    Returns:
        Nested dict: {sheet_name: {cell_coord: value}}
    """
    sheets_data = {}
    current_sheet = None
    current_cell_value = None  # For multi-line handling
    current_cell_coord = None
    
    lines = csv_content.split('\n')
    
    for line in lines:
        # Don't strip - preserve whitespace in sheet names
        
        # Detect sheet header (keep original whitespace in sheet name)
        if line.strip().startswith('Sheet:'):
            # Extract sheet name, preserving any trailing spaces
            current_sheet = line.strip().replace('Sheet:', '', 1).strip()
            sheets_data[current_sheet] = {}
            continue
        
        # Skip separator lines
        if line.strip().startswith('=') or not line.strip():
            continue
        
        # Parse cell data: "A1: value,B1: value2,C1:"
        if current_sheet is not None:
            # Split on comma followed by cell coord, but NOT on escaped commas (\,)
            cells = re.split(r'(?<!\\),(?=[A-Z]+\d+:)', line)
            
            for cell in cells:
                cell = cell.strip()
                if not cell:
                    continue
                    
                match = re.match(r'^([A-Z]+\d+):\s*(.*)', cell)
                if match:
                    coord = match.group(1)
                    value = match.group(2)
                    value = value.replace('\\,', ',')  # Unescape commas
                    sheets_data[current_sheet][coord] = value
    
    return sheets_data


def get_excel_cell_values(excel_path: str, sheet_name: str) -> Dict[str, str]:
    """Extract all cell values from an Excel sheet."""
    wb = load_workbook(excel_path, data_only=True)
    sheet = wb[sheet_name]
    
    cell_values = {}
    
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, max_col=sheet.max_column):
        for cell in row:
            coord = f"{get_column_letter(cell.column)}{cell.row}"
            if cell.value is not None:
                cell_values[coord] = str(cell.value)
            else:
                cell_values[coord] = ""
    
    wb.close()
    return cell_values


def normalize_value(value: str) -> str:
    """Normalize a value for comparison."""
    if value is None:
        return ""
    value = str(value).strip()
    value = re.sub(r'\s+', ' ', value)
    return value


def is_image_reference(value: str) -> bool:
    """Check if a value is a markdown image reference."""
    return value.startswith('![') and '](images/' in value


def compare_sheet(
    excel_path: str,
    sheet_name: str,
    csv_cells: Dict[str, str]
) -> SheetMetrics:
    """Compare a single sheet's CSV output against the original Excel."""
    metrics = SheetMetrics(sheet_name=sheet_name)
    
    # Try exact match first, then try with/without trailing space
    wb = load_workbook(excel_path, data_only=True)
    
    actual_sheet_name = None
    if sheet_name in wb.sheetnames:
        actual_sheet_name = sheet_name
    else:
        # Try fuzzy match (trailing/leading whitespace)
        for name in wb.sheetnames:
            if name.strip() == sheet_name.strip():
                actual_sheet_name = name
                break
    
    if actual_sheet_name is None:
        wb.close()
        raise ValueError(f"Worksheet {sheet_name} does not exist.")
    
    sheet = wb[actual_sheet_name]
    
    excel_cells = {}
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, max_col=sheet.max_column):
        for cell in row:
            coord = f"{get_column_letter(cell.column)}{cell.row}"
            if cell.value is not None:
                value = str(cell.value)
                value = value.replace('\n', '\\n')  # Match CSV format
                excel_cells[coord] = value
            else:
                excel_cells[coord] = ""
    
    wb.close()
    
    all_coords = set(excel_cells.keys()) | set(csv_cells.keys())
    
    for coord in all_coords:
        excel_value = normalize_value(excel_cells.get(coord, ""))
        csv_value = normalize_value(csv_cells.get(coord, ""))
        
        # Skip empty cells in both
        if not excel_value and not csv_value:
            continue
        
        # Count image references separately
        if is_image_reference(csv_value):
            metrics.images_found += 1
            continue
        
        metrics.total_cells += 1
        
        if excel_value == csv_value:
            metrics.matched_cells += 1
        else:
            metrics.mismatched_cells += 1
            
            if len(metrics.mismatches) < 10:
                metrics.mismatches.append({
                    'cell': coord,
                    'excel': excel_value[:50] + ('...' if len(excel_value) > 50 else ''),
                    'csv': csv_value[:50] + ('...' if len(csv_value) > 50 else ''),
                })
    
    return metrics


def run_and_measure(excel_path: str, output_path: str) -> OverallMetrics:
    """
    Run the full conversion pipeline and measure everything.
    
    1. Deletes any existing output (clean slate)
    2. Runs excel_converter.process_all_sheets() with timing
    3. Reads the output CSV
    4. Compares against original Excel for accuracy
    
    Args:
        excel_path: Path to input Excel file
        output_path: Path for output CSV
        
    Returns:
        OverallMetrics with latency and accuracy data
    """
    overall = OverallMetrics()
    
    # ==================== CLEAN UP EXISTING OUTPUT ====================
    if os.path.exists(output_path):
        os.remove(output_path)
        print(f"🗑️  Removed existing: {output_path}")
    
    images_dir = os.path.join(os.path.dirname(output_path) or '.', 'images')
    if os.path.exists(images_dir):
        for f in os.listdir(images_dir):
            os.remove(os.path.join(images_dir, f))
        print(f"🗑️  Cleared existing images/")
    # ==================================================================
    
    print("\n" + "=" * 80)
    print("STEP 1: RUNNING CONVERSION (measuring latency)")
    print("=" * 80 + "\n")
    
    # ==================== MEASURE CONVERSION LATENCY ====================
    conversion_start = time.time()
    
    process_all_sheets(excel_path, output_path)
    
    overall.conversion_time = time.time() - conversion_start
    # ====================================================================
    
    print(f"\n✓ Conversion completed in {overall.conversion_time:.2f}s")
    
    # Verify output was created
    if not os.path.exists(output_path):
        print(f"✗ ERROR: Output file not created: {output_path}")
        return overall
    
    print("\n" + "=" * 80)
    print("STEP 2: MEASURING ACCURACY")
    print("=" * 80 + "\n")
    
    # ==================== MEASURE ACCURACY ====================
    accuracy_start = time.time()
    
    # Read the generated CSV
    with open(output_path, 'r', encoding='utf-8') as f:
        csv_content = f.read()
    
    # Parse CSV
    csv_data = parse_csv_output(csv_content)
    print(f"Parsed {len(csv_data)} sheets from CSV output\n")
    
    # Compare each sheet
    for sheet_name, csv_cells in csv_data.items():
        try:
            metrics = compare_sheet(excel_path, sheet_name, csv_cells)
            overall.sheets.append(metrics)
            
            status = "✓" if metrics.accuracy == 100.0 else "✗"
            print(f"{status} '{sheet_name}': {metrics.accuracy:.1f}% "
                  f"({metrics.matched_cells}/{metrics.total_cells} cells, "
                  f"{metrics.images_found} images)")
            
            if metrics.mismatches:
                for m in metrics.mismatches[:3]:
                    print(f"    {m['cell']}: Excel='{m['excel']}' vs CSV='{m['csv']}'")
                if len(metrics.mismatches) > 3:
                    print(f"    ... and {len(metrics.mismatches) - 3} more mismatches")
                    
        except Exception as e:
            print(f"✗ '{sheet_name}': Error - {e}")
    
    overall.accuracy_check_time = time.time() - accuracy_start
    # ==========================================================
    
    return overall


def print_summary(metrics: OverallMetrics, excel_path: str):
    """Print final summary."""
    sheet_count = len(get_all_sheet_names(excel_path))
    
    print("\n" + "=" * 80)
    print("SUMMARY")
    print("=" * 80)
    
    print(f"\n📊 ACCURACY:")
    print(f"   Overall: {metrics.overall_accuracy:.2f}%")
    print(f"   Cells compared: {metrics.total_cells}")
    print(f"   Matched: {metrics.total_matched}")
    print(f"   Mismatched: {metrics.total_cells - metrics.total_matched}")
    print(f"   Images found: {metrics.total_images}")
    
    print(f"\n⏱️  LATENCY:")
    print(f"   Conversion time: {metrics.conversion_time:.2f}s")
    print(f"   Sheets processed: {sheet_count}")
    print(f"   Avg per sheet: {metrics.conversion_time / sheet_count:.3f}s")
    print(f"   Accuracy check time: {metrics.accuracy_check_time:.2f}s")
    
    # Worst sheets
    if metrics.sheets:
        worst = sorted(metrics.sheets, key=lambda s: s.accuracy)[:3]
        if worst and worst[0].accuracy < 100:
            print(f"\n⚠️  LOWEST ACCURACY SHEETS:")
            for s in worst:
                if s.accuracy < 100:
                    print(f"   {s.sheet_name}: {s.accuracy:.1f}%")
    
    print("\n" + "=" * 80)


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    
    excel_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else 'output_measured.csv'
    
    print(f"\n🔍 Input:  {excel_file}")
    print(f"📄 Output: {output_file}")
    
    metrics = run_and_measure(excel_file, output_file)
    print_summary(metrics, excel_file)